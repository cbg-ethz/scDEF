"""Download 10x Genomics CytAssist FFPE Human Breast Cancer Visium data
and matched Chromium scFFPE-seq reference (Janesick et al. 2023, Nature Communications).

Visium spatial data: publicly available from cf.10xgenomics.com.
Chromium scFFPE-seq + cell type annotations: from GEO (GSE243275 / GSM7782698).
"""
import os
import urllib.request
import tarfile
import gzip
import shutil

VISIUM_H5_URL = "https://cf.10xgenomics.com/samples/spatial-exp/2.0.0/CytAssist_FFPE_Human_Breast_Cancer/CytAssist_FFPE_Human_Breast_Cancer_filtered_feature_bc_matrix.h5"
VISIUM_SPATIAL_URL = "https://cf.10xgenomics.com/samples/spatial-exp/2.0.0/CytAssist_FFPE_Human_Breast_Cancer/CytAssist_FFPE_Human_Breast_Cancer_spatial.tar.gz"

GEO_SCRNA_URL = "https://ftp.ncbi.nlm.nih.gov/geo/samples/GSM7782nnn/GSM7782698/suppl/GSM7782698_count_raw_feature_bc_matrix.h5"
GEO_ANNOT_URL = "https://ftp.ncbi.nlm.nih.gov/geo/series/GSE243nnn/GSE243275/suppl/GSE243275_Barcode_Cell_Type_Matrices.xlsx"


def _download(url, dest):
    """Download a file with browser-like headers to avoid 403 from CDNs."""
    print(f"Downloading {url} ...")
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/120.0.0.0 Safari/537.36",
        })
        with urllib.request.urlopen(req) as response:
            with open(dest, "wb") as out_file:
                shutil.copyfileobj(response, out_file)
        print(f"  -> {dest} ({os.path.getsize(dest)} bytes)")
        return True
    except Exception as e:
        print(f"  WARNING: download failed: {e}")
        return False


def _extract_celltypes_csv(xlsx_path, csv_path):
    """Extract scFFPE-seq cell type annotations from the Barcode_Cell_Type_Matrices xlsx."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        # Find the scFFPE-seq / Flex sheet
        target_sheet = None
        for name in wb.sheetnames:
            if "flex" in name.lower() or "scffpe" in name.lower() or "chromium" in name.lower():
                target_sheet = name
                break
        if target_sheet is None:
            # Fall back to first sheet that isn't Xenium or Visium
            for name in wb.sheetnames:
                if "xenium" not in name.lower() and "visium" not in name.lower():
                    target_sheet = name
                    break
        if target_sheet is None:
            target_sheet = wb.sheetnames[0]

        print(f"  Extracting celltypes from sheet: '{target_sheet}'")
        ws = wb[target_sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        import csv
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            for row in rows[1:]:
                writer.writerow(row)
        wb.close()
        print(f"  -> {csv_path}")
        return True
    except ImportError:
        print("  WARNING: openpyxl not installed, cannot extract celltypes.csv from xlsx")
        print(f"  Please manually open {xlsx_path} and save the scFFPE-seq tab as {csv_path}")
        return False
    except Exception as e:
        print(f"  WARNING: failed to extract celltypes: {e}")
        return False


def main():
    out_dir = snakemake.params["out_dir"]
    os.makedirs(out_dir, exist_ok=True)

    # --- Visium spatial data ---
    h5_path = os.path.join(out_dir, "CytAssist_FFPE_Human_Breast_Cancer_filtered_feature_bc_matrix.h5")
    if not _download(VISIUM_H5_URL, h5_path):
        raise RuntimeError("Failed to download Visium h5 matrix")

    tar_path = os.path.join(out_dir, "spatial.tar.gz")
    if _download(VISIUM_SPATIAL_URL, tar_path):
        with tarfile.open(tar_path, "r:gz") as tar:
            tar.extractall(path=out_dir)
        os.remove(tar_path)
    else:
        raise RuntimeError("Failed to download Visium spatial data")

    # --- Matched Chromium scFFPE-seq reference (from GEO) ---
    scrna_path = os.path.join(out_dir, "Chromium_FFPE_Human_Breast_Cancer_filtered_feature_bc_matrix.h5")
    if not os.path.exists(scrna_path):
        _download(GEO_SCRNA_URL, scrna_path)

    # --- Cell type annotations (from GEO) ---
    xlsx_path = os.path.join(out_dir, "Barcode_Cell_Type_Matrices.xlsx")
    csv_path = os.path.join(out_dir, "celltypes.csv")
    if not os.path.exists(csv_path):
        if not os.path.exists(xlsx_path):
            _download(GEO_ANNOT_URL, xlsx_path)
        _extract_celltypes_csv(xlsx_path, csv_path)

    with open(snakemake.output["done"], "w") as f:
        f.write("done\n")

    print("Done.")


if __name__ == "__main__":
    main()
